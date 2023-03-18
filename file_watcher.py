from os import listdir
from os.path import isfile, join

import os
import pathlib
import shutil
import sys
import time
import pandas as pd
import openpyxl as xl;

watchDirectory = r"C:\Users\juanc\OneDrive\Documents\Projects\programming_excersise\files"
processedPath = r'C:\Users\juanc\OneDrive\Documents\Projects\programming_excersise\files\Processed'
notApplicablePath= r'C:\Users\juanc\OneDrive\Documents\Projects\programming_excersise\files\Not Applicable'
master_file =r"C:\Users\juanc\OneDrive\Documents\Projects\programming_excersise\master.xlsx"

#The interval at which we are going to review if any file was added. 
pollTime = 2 

"""
    Returns all the files in a given directory.
"""
def get_files_dir(folder_path: str) -> list:
    only_files = [f for f in listdir(folder_path) if isfile(join(folder_path, f))]
    return only_files

"""
    Function that determines if there are diferences between two lists.
"""
def compare_lists(original_list: list, new_list: list) -> list:
    diff_list = [x for x in new_list if x not in original_list]
    return diff_list

"""
    Once the watcher determines that we have new files to work with, we work with them.
"""
def new_file_actions(newFiles: list, previousFileList: list) -> list:
    for x in newFiles:
        if str(get_file_ext(x)).startswith(".xls"):
            shutil.copyfile(os.path.join(watchDirectory, x), os.path.join(processedPath, x))
            path_new_file=os.path.join(watchDirectory, x)
            consolidate_master(path_new_file)
        else: 
            shutil.copyfile(os.path.join(watchDirectory, x), os.path.join(notApplicablePath, x))
        previousFileList.append(x)

    return previousFileList
    
"""
    Watches a directory every X amount of time, where X is the variable pollTime. And process the information
"""
def file_watcher(my_dir: str, pollTime: int):
    workbook = xl.Workbook()
    worksheet=workbook.active
    workbook.save(str(master_file))
    
    while True:
        if 'watching' not in locals(): #Check if this is the first time the function has run
            previousFileList = get_files_dir(watchDirectory)
            watching = 1
            print('First Time')
            print(previousFileList)
        
        time.sleep(pollTime)
        newFileList = get_files_dir(watchDirectory)
        fileDiff = compare_lists(previousFileList, newFileList)
        previousFileList = new_file_actions(fileDiff, previousFileList)

"""
    Saves all the sheets from found files into the main file.
"""
def consolidate_master(path_new_file: str):
    # opening the source excel file
    filename = path_new_file
    wb_newfile = xl.load_workbook(filename)
    for sheet in wb_newfile.worksheets:
        wb_newsheet = sheet
        # Opening the destination excel file
        wb_master = xl.load_workbook(master_file)
        wb_master.create_sheet(sheet.title)
        ws_master = wb_master[sheet.title]

        # calculate total number of rows and
        # columns in source excel file
        mr = wb_newsheet.max_row
        mc = wb_newsheet.max_column

        # copying the cell values from source
        # excel file to destination excel file
        for i in range (1, mr + 1):
            for j in range (1, mc + 1):
                # reading cell value from source excel file
                c = wb_newsheet.cell(row = i, column = j)
                # writing the read value to destination excel file
                ws_master.cell(row = i, column = j).value = c.value

        # saving the destination excel file
        ws_master.save(str(master_file))

"""
    Returns the extension of the file that was sent.
"""
def get_file_ext(path: str) -> str:
    extension = pathlib.Path(path).suffix
    return extension
    
if __name__ == "__main__":
    file_watcher(watchDirectory, pollTime)